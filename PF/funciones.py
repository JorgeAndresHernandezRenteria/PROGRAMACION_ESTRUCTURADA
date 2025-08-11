import os
from mysql.connector import Error
from conexionBD import conectar
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from datetime import datetime
import re

def borrar_pantalla():
    os.system('cls' if os.name == 'nt' else 'clear')

def esperar_tecla():
    input("\nPresione cualquier tecla para continuar...")

def menu_principal():
    borrar_pantalla()
    print("\n" + "=" * 50)
    print(" SISTEMA GESTIÓN SIX ")
    print("=" * 50)
    print("1. Iniciar sesión")
    print("2. Registrar usuario")
    print("3. Salir")
    return input("\nSeleccione una opción: ")

def menu_inventario():
    borrar_pantalla()
    print("\n" + "=" * 50)
    print(" MENÚ PRINCIPAL - GESTIÓN SIX")
    print("=" * 50)
    print("1. Realizar venta")
    print("2. Registrar producto")
    print("3. Listar productos")
    print("4. Buscar producto")
    print("5. Actualizar stock")
    print("6. Menú de administración de tablas")
    print("7. Cerrar sesión")
    return input("\nSeleccione una opción: ")

def menu_admin():
    borrar_pantalla()
    print("\n" + "=" * 50)
    print(" MENÚ DE ADMINISTRACIÓN DE TABLAS")
    print("=" * 50)
    print("1. Reiniciar tabla de productos")
    print("2. Reiniciar tabla de ventas")
    print("3. Reiniciar tabla de detalle de ventas")
    print("4. Reiniciar tabla de usuarios")
    print("5. Exportar tablas a PDF y/ó Excel...")
    print("6. Volver al menú principal")
    return input("\nSeleccione una opción: ")

def formatear_precio(precio):
    return f"${precio:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def tabla_excel(datos, nombre_archivo="reporte"):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte del Sistema SIX"
        
        # Encabezados
        encabezados = list(datos[0].keys()) if datos else []
        for col, encabezado in enumerate(encabezados, start=1):
            ws.cell(row=1, column=col, value=encabezado).font = Font(bold=True)
        
        # Datos
        for row, item in enumerate(datos, start=2):
            for col, (key, value) in enumerate(item.items(), start=1):
                ws.cell(row=row, column=col, value=value)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
        
        nombre_completo = f"{nombre_archivo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(nombre_completo)
        return nombre_completo
    except Exception as e:
        print(f"Error al exportar a Excel: {e}")
        return None

def tabla_pdf(datos, nombre_archivo="reporte"):
    try:
        # Configuración del nombre de archivo
        nombre_archivo = f"{re.sub(r'[^\w-]', '', nombre_archivo)}.pdf"
        
        doc = SimpleDocTemplate(
            nombre_archivo,
            pagesize=letter,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Título limpio
        elements.append(Paragraph("Reporte del Sistema SIX", styles['Title']))
        elements.append(Spacer(1, 12))
        
        # Limpieza de datos
        datos_limpios = []
        encabezados = []
        
        if datos and len(datos) > 0:
            encabezados = [str(key) for key in datos[0].keys()]
            
            for item in datos:
                fila = []
                for key in encabezados:
                    valor = str(item.get(key, '')).replace('\n', ' ').replace('\r', '')
                    fila.append(valor[:100])  # Limitar a 100 caracteres
                datos_limpios.append(fila)
        
        # Crear tabla
        tabla_datos = [encabezados] + datos_limpios
        
        # Estilo 
        estilo = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0066CC')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F2F2F2')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
        ])
        
        tabla = Table(tabla_datos)
        tabla.setStyle(estilo)
        elements.append(tabla)
        
        # Pie de página 
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(
            f"Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            styles['Italic']
        ))
        
        doc.build(elements)
        return nombre_archivo
        
    except Exception as e:
        print(f"Error al generar PDF: {str(e)[:200]}")
        return None
    
def exportar_tabla(tabla, formato):
    #Exporta una tabla específica al formato seleccionado

    # Obtener datos de la tabla
    datos = obtener_datos_tabla(tabla)
    if not datos:
        return None

    nombre_archivo = f"reporte{tabla} {datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    if formato == 'excel':
        return tabla_excel(datos, nombre_archivo)
    elif formato == 'pdf':
        return tabla_pdf(datos, nombre_archivo)
    return None

def obtener_datos_tabla(tabla):
    """Obtiene todos los datos de una tabla específica"""
    conexion = conectar()
    if not conexion:
        return None

    try:
        cursor = conexion.cursor(dictionary=True)
        cursor.execute(f"SELECT * FROM {tabla}")
        return cursor.fetchall()
    except Error as e:
        print(f"Error al obtener datos de {tabla}: {e}")
        return None
    finally:
        if conexion.is_connected():
            conexion.close()